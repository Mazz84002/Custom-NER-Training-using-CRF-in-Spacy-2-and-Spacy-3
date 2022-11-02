# converting csv type into .txt like other datasets
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook(filename='/Users/mazz/Documents/Programming/Python/Mini-Project/DataSet/ner_dataset.xlsx')
list_of_sheets = wb.sheetnames
sh = wb[list_of_sheets[0]]
train = open('/Users/mazz/Documents/Programming/Python/Mini-Project/DataSet/dataset3_train.txt', "w+")
test = open('/Users/mazz/Documents/Programming/Python/Mini-Project/DataSet/dataset3_test.txt', "w+")
train.write("-DOCSTART- -X- -X- O" + "\n")
test.write("-DOCSTART- -X- -X- O" + "\n")

n = sh.max_row+1

for i in range(1, int(0.8*n)):
    if str(sh.cell(row=i, column=1).value)[0:8] == 'Sentence':
        train.write("\n")
        train.write(str(sh.cell(row=i, column=2).value) + " " + str(sh.cell(row=i, column=3).value) + " " + str(sh.cell(row=i, column=4).value) + "\n")
    else:
        train.write(str(sh.cell(row=i, column=2).value) + " " + str(sh.cell(row=i, column=3).value) + " " + str(sh.cell(row=i, column=4).value) + "\n")

for i in range(int(0.8*n), n):
    if str(sh.cell(row=i, column=1).value)[0:8] == 'Sentence':
        test.write("\n")
        test.write(str(sh.cell(row=i, column=2).value) + " " + str(sh.cell(row=i, column=3).value) + " " + str(sh.cell(row=i, column=4).value) + "\n")
    else:
        test.write(str(sh.cell(row=i, column=2).value) + " " + str(sh.cell(row=i, column=3).value) + " " + str(sh.cell(row=i, column=4).value) + "\n")

train.close()
test.close()
wb.close()
